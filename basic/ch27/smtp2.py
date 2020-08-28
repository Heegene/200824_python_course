# mail message 사용 객체 선언

from email.message import EmailMessage
import smtplib
import openpyxl


def check_party_members():
    # Excel open
    wb = openpyxl.load_workbook("inv.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    party_members={}

    # 2 row 부터 max row 까지 진행
    for row in range(2, sheet.max_row + 1):
        payment = sheet.cell(row=row, column=sheet.max_column).value
        if payment == 'invite':
            name = sheet.cell(row=row, column=1).value
            email = sheet.cell(row=row, column=2).value
            party_members[name] = email
    return party_members


def main():
    # 명단
    party_members = check_party_members()
    print('main party members {}...'.format(party_members))
    # 로그인
    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login('dev.heejin@gmail.com', '앱 비밀번호')

    for name, email in party_members.items():
        # 메세지 생성
        msg = EmailMessage()

        # 메시지 내용 지정
        msg['Subject'] = '회원 파티 초대장(mail test)'
        msg['From'] = 'dev.heejin@gmail.com'
        msg['to'] = email
        msg.set_content(''' 안녕하세요 {} 님,
                            2020년 08월 저녁파티에 참석 요청 합니다.
                            빠른 시간 내 회신 부탁드려요. '''.format(name))
        print("sending email to {}".format(email))
        sendmail_status = smtp.send_message(msg)
        # 메일전송

        if sendmail_status != "":
            print("{}에게 메일 전송이 완료되었습니다.".format(email, sendmail_status))
        else:
            print("{}에게 메일 전송이 실패하였습니다.".format(email, sendmail_status))
    smtp.quit()


if __name__ == '__main__':
    main()

